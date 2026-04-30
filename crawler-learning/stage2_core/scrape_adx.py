import requests
from bs4 import BeautifulSoup
import openpyxl
import time
import random


def scrape_all_exhibitors():
    """抓取所有展商（支持分页加载）"""
    base_url = "https://www.adx.sydney"
    list_url = f"{base_url}/exhibitor-list.html"
    ajax_url = f"{base_url}/exhibitor/index/load-more/perPage/20/j/62a40feb6b9db466788a96bc617bc37e"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": list_url,
    }

    all_exhibitors = []

    print("正在获取第1页...")
    resp = requests.get(list_url, headers=headers, timeout=30)
    resp.encoding = "utf-8"
    soup = BeautifulSoup(resp.text, "lxml")

    # 解析第1页
    for a in soup.select("a.exhibitor-link"):
        name = a.text.strip()
        exhibitor_link = a.get("data-exhibitorlink", "")
        if exhibitor_link:
            all_exhibitors.append(
                {"name": name, "ajax_url": f"{base_url}{exhibitor_link}"}
            )

    print(f"第1页找到 {len(all_exhibitors)} 个展商")

    # 翻页加载更多
    page = 2
    max_pages = 20  # 安全限制

    while page <= max_pages:
        print(f"正在加载第 {page} 页...")

        post_data = {
            "page": page,
            "showSearch": "",
            "businessName-autocomplete": "",
            "cateogryId": "",
            "showThumbnails": 0,
        }

        try:
            resp = requests.post(ajax_url, data=post_data, headers=headers, timeout=30)

            if resp.status_code != 200 or not resp.text.strip():
                print(f"第 {page} 页为空或请求失败，停止")
                break

            soup = BeautifulSoup(resp.text, "lxml")

            # 解析这页的展商
            page_exhibitors = []
            for a in soup.select("a.exhibitor-link"):
                name = a.text.strip()
                exhibitor_link = a.get("data-exhibitorlink", "")
                if exhibitor_link:
                    page_exhibitors.append(
                        {"name": name, "ajax_url": f"{base_url}{exhibitor_link}"}
                    )

            if not page_exhibitors:
                print(f"第 {page} 页没有展商，停止")
                break

            all_exhibitors.extend(page_exhibitors)
            print(
                f"第 {page} 页找到 {len(page_exhibitors)} 个，累计 {len(all_exhibitors)} 个"
            )

            # 如果返回的展商数少于20，可能是最后一页
            if len(page_exhibitors) < 20:
                print("已加载完所有展商")
                break

        except Exception as e:
            print(f"第 {page} 页加载失败: {e}")
            break

        page += 1
        time.sleep(random.uniform(0.3, 0.7))

    print(f"\n共找到 {len(all_exhibitors)} 个展商")
    return all_exhibitors


def scrape_ajax_detail(ajax_url, headers):
    """请求AJAX接口获取展商详情"""
    try:
        resp = requests.get(ajax_url, headers=headers, timeout=30)
        resp.encoding = "utf-8"

        if resp.status_code != 200:
            return None

        soup = BeautifulSoup(resp.text, "lxml")

        table = soup.select_one("#ivt-exhibitor-details")

        data = {"booths": "", "address": "", "phone": "", "email": "", "website": ""}

        if table:
            for row in table.select("tr"):
                label = row.select_one(".label")
                value = row.select_one(".value")
                if label and value:
                    label_text = label.text.strip().lower()
                    value_text = value.text.strip()

                    value_link = value.select_one("a")
                    value_href = value_link.get("href", "") if value_link else ""

                    if "booth" in label_text:
                        data["booths"] = value_text
                    elif "address" in label_text:
                        data["address"] = value_text
                    elif "phone" in label_text:
                        data["phone"] = value_text
                    elif "email" in label_text:
                        if value_link and "@" in value_link.text:
                            data["email"] = value_link.text.strip()
                        else:
                            data["email"] = value_text
                    elif "website" in label_text:
                        if value_href and value_href != "http://":
                            data["website"] = value_href

        return data
    except Exception as e:
        print(f"Error fetching {ajax_url}: {e}")
        return None


def main():
    exhibitors = scrape_all_exhibitors()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ADX Sydney Exhibitors"

    headers_list = ["公司名称", "AJAX链接", "Booths", "地址", "电话", "邮箱", "网站"]
    for col, h in enumerate(headers_list, 1):
        ws.cell(row=1, column=col, value=h)

    request_headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": "https://www.adx.sydney/exhibitor-list.html",
    }

    for i, ex in enumerate(exhibitors):
        ws.cell(row=i + 2, column=1, value=ex["name"])
        ws.cell(row=i + 2, column=2, value=ex["ajax_url"])

        detail = scrape_ajax_detail(ex["ajax_url"], request_headers)

        if detail:
            ws.cell(row=i + 2, column=3, value=detail["booths"])
            ws.cell(row=i + 2, column=4, value=detail["address"])
            ws.cell(row=i + 2, column=5, value=detail["phone"])
            ws.cell(row=i + 2, column=6, value=detail["email"])
            ws.cell(row=i + 2, column=7, value=detail["website"])

            if (i + 1) % 10 == 0:
                print(f"[{i + 1}/{len(exhibitors)}] {ex['name']} - {detail['email']}")
        else:
            print(f"[{i + 1}/{len(exhibitors)}] Failed: {ex['name']}")

        time.sleep(random.uniform(0.3, 0.7))

        if (i + 1) % 50 == 0:
            wb.save("adx_exhibitors.xlsx")
            print(f"已保存 {i + 1}/{len(exhibitors)}")

    wb.save("adx_exhibitors.xlsx")
    print(f"\n完成! 共处理 {len(exhibitors)} 个展商，保存到 adx_exhibitors.xlsx")


if __name__ == "__main__":
    main()
