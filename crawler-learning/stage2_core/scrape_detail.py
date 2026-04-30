import cloudscraper
from bs4 import BeautifulSoup
import openpyxl
import time


def scrape_detail(url, scraper):
    try:
        resp = scraper.get(url, timeout=30)
        resp.encoding = "utf-8"
        if resp.status_code != 200:
            return None, None, None, None

        soup = BeautifulSoup(resp.text, "lxml")

        name = soup.select_one("h1.nocap")
        name = name.text.strip() if name else ""

        address = ""
        website = ""

        gc_six = soup.select_one("div.gridcol.six")
        if gc_six:
            a_tag = gc_six.select_one("a")
            if a_tag:
                href = a_tag.get("href", "")
                if href.startswith("http") and "mapyourshow" not in href:
                    website = href
            for a in gc_six.select("a"):
                a.decompose()
            address = gc_six.get_text(separator=" ").strip()
            address = " ".join(address.split())

        if not address:
            for gc in soup.select("div.gridcol.four"):
                links = gc.select("a")
                for link in links:
                    href = link.get("href", "")
                    if href.startswith("http") and "mapyourshow" not in href:
                        if not any(
                            x in href
                            for x in [
                                "facebook",
                                "twitter",
                                "youtube",
                                "instagram",
                                "linkedin",
                                "tiktok",
                                "automate.org",
                            ]
                        ):
                            website = href
                            text = gc.get_text(separator=" ").strip()
                            address = text.replace(website, "").strip()
                            address = " ".join(address.split())
                            break
                if address:
                    break

        desc_div = soup.select_one("div.exhibitor-description")
        description = ""
        if desc_div:
            paragraphs = desc_div.select("p")
            description = " ".join(p.text.strip() for p in paragraphs if p.text.strip())

        return name, address, website, description
    except Exception as e:
        print(f"Error: {url} - {e}")
        return None, None, None, None


def main():
    print("读取现有Excel获取展商链接...")
    wb = openpyxl.load_workbook("exhibitors_full.xlsx")
    ws = wb.active

    links = []
    for row in range(2, ws.max_row + 1):
        link = ws.cell(row=row, column=1).value
        if link:
            links.append((row, link))

    print(f"共有 {len(links)} 个展商链接")

    # 创建云爬虫
    scraper = cloudscraper.create_scraper()

    # 创建新的 Excel 文件
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "展商详情"

    # 设置表头
    headers = ["链接", "原名称", "展位号", "公司名称", "地址", "官网", "描述"]
    for col, h in enumerate(headers, 1):
        new_ws.cell(row=1, column=col, value=h)  # 设置表头

    # 读取原始 Excel 文件
    original_wb = openpyxl.load_workbook("exhibitors_full.xlsx")
    original_ws = original_wb.active

    # 遍历展商链接
    for idx, (original_row, link) in enumerate(links):
        # 获取原始名称和展位号
        name = original_ws.cell(row=original_row, column=2).value or ""
        booth = original_ws.cell(row=original_row, column=3).value or ""

        # 设置新的 Excel 文件的表头
        new_ws.cell(row=idx + 2, column=1, value=link)
        new_ws.cell(row=idx + 2, column=2, value=name)
        new_ws.cell(row=idx + 2, column=3, value=booth)

        # 获取展商详情
        detail_name, address, website, description = scrape_detail(link, scraper)

        if detail_name:
            # 设置新的 Excel 文件的表头
            new_ws.cell(row=idx + 2, column=4, value=detail_name)
            new_ws.cell(row=idx + 2, column=5, value=address)  # 设置地址
            new_ws.cell(row=idx + 2, column=6, value=website)  # 设置官网
            new_ws.cell(row=idx + 2, column=7, value=description)  # 设置描述
            if (idx + 1) % 10 == 0:
                print(f"[{idx + 1}/{len(links)}] {detail_name}")
        else:
            print(f"[{idx + 1}/{len(links)}] Failed: {link}")

        if (idx + 1) % 50 == 0:
            print(f"已处理 {idx + 1}/{len(links)}，保存进度...")
            new_wb.save("exhibitors_detail.xlsx")

        time.sleep(0.5)

    new_wb.save("exhibitors_detail.xlsx")
    print(f"\n完成! 共处理 {len(links)} 个展商，保存到 exhibitors_detail.xlsx")


if __name__ == "__main__":
    main()
